#!/usr/bin/env python3
"""
FACS Autogating - Style FlowJo
Workflow immunophénotypage complet avec:
- Gating hiérarchique (Cells → Single Cells → Live → hCD45+ → populations)
- Quadrants pour sous-populations (T cells, B cells, NK cells, Treg)
- Histogrammes pour marqueurs fonctionnels
- Visualisation style FlowJo (pseudocolor, gates rectangulaires)
"""

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.colors import LogNorm
from pathlib import Path
import tempfile
import io
from datetime import datetime
import flowio
import re

# Configuration
st.set_page_config(
    page_title="FACS - Style FlowJo",
    page_icon="🔬",
    layout="wide"
)

# CSS
st.markdown("""
    <style>
    .main-header { font-size: 2.2rem; color: #2c3e50; text-align: center; margin-bottom: 1rem; }
    .population-box { background: #f8f9fa; padding: 0.5rem; border-radius: 0.3rem; 
                      border-left: 3px solid #3498db; margin: 0.3rem 0; font-size: 0.9rem; }
    .stats-table { font-size: 0.85rem; }
    </style>
""", unsafe_allow_html=True)


# ==================== CHANNEL MAPPING ====================
CHANNEL_MARKERS = {
    'FSC-A': 'FSC-A',
    'FSC-H': 'FSC-H', 
    'SSC-A': 'SSC-A',
    'BUV395-A': 'PDL1',
    'BUV805-A': 'CD8',
    'eFluor450-A': 'FoxP3',
    'LiveDeadFixableAqua-A': 'Live/Dead',
    'BV650-A': 'CD4',
    'BV711-A': 'CD161',
    'BV785-A': 'CD25',
    '[RB780]-A': 'PD1',
    'AF488-A': 'CD3',
    'NovaFluorBlue585-A': 'CD16',
    'PerCP-Vio700-A': 'CD107a',
    'PerCP-A': 'hCD45',
    'PE-A': 'LLT1',
    'PE-Dazzle594-A': 'Granzyme B',
    'PE-Fire700-A': 'CD19',
    'PE-Cy7-A': 'CD56',
    'APC-A': 'HLA-ABC',
    'APC-Fire750-A': 'mCD45',
}


class FCSReader:
    def __init__(self, fcs_path):
        self.fcs_path = fcs_path
        self.flow_data = flowio.FlowData(fcs_path)
        self.data = None
        self.channels = []
        self.channel_info = {}
        self.filename = Path(fcs_path).stem
        self.load_data()
    
    def load_data(self):
        events = self.flow_data.events
        n_channels = self.flow_data.channel_count
        
        if not isinstance(events, np.ndarray):
            events = np.array(events, dtype=np.float64)
        
        if events.ndim == 1:
            n_events = len(events) // n_channels
            events = events.reshape(n_events, n_channels)
        
        pnn_labels = []
        for i in range(1, n_channels + 1):
            pnn = self.flow_data.text.get(f'$P{i}N', None) or self.flow_data.text.get(f'p{i}n', f'Ch{i}')
            pns = self.flow_data.text.get(f'$P{i}S', None) or self.flow_data.text.get(f'p{i}s', '')
            pnn = pnn.strip() if pnn else f'Ch{i}'
            pns = pns.strip() if pns else ''
            
            # Extraire le marqueur
            marker = CHANNEL_MARKERS.get(pnn, None)
            if not marker and pns:
                # Extraire depuis PnS (ex: "hCD4 : BV650 - Area" -> CD4)
                match = re.search(r'[hm]?(CD\d+[a-z]?|FoxP3|Granzyme|PD[L]?1|HLA)', pns, re.IGNORECASE)
                if match:
                    marker = match.group(0).replace('h', '').replace('m', '')
            
            self.channel_info[pnn] = {'pns': pns, 'marker': marker or pnn}
            pnn_labels.append(pnn)
        
        self.channels = pnn_labels
        self.data = pd.DataFrame(events, columns=self.channels)
        return self.data
    
    def get_marker(self, channel):
        return self.channel_info.get(channel, {}).get('marker', channel)


def find_channel(data, keywords):
    """Trouve un canal par mots-clés"""
    for col in data.columns:
        col_upper = col.upper()
        for kw in keywords:
            if kw.upper() in col_upper:
                return col
    return None


def biex_transform(x, width=5, negative=0):
    """Transformation biexponentielle simplifiée (style FlowJo)"""
    x = np.asarray(x, dtype=float)
    # Transformation asinh pour approximer biexponentielle
    return np.arcsinh(x / 150) * 50


def create_flowjo_plot(ax, x_data, y_data, x_label, y_label, title="", 
                       gate_coords=None, gate_label=None, gate_pct=None,
                       quadrant_coords=None, quadrant_labels=None, quadrant_pcts=None,
                       show_stats=True, cmap='jet'):
    """Crée un plot style FlowJo avec pseudocolor"""
    
    # Filtrer données valides
    valid = np.isfinite(x_data) & np.isfinite(y_data)
    x_plot = x_data[valid]
    y_plot = y_data[valid]
    
    if len(x_plot) == 0:
        ax.text(0.5, 0.5, 'No data', ha='center', va='center', transform=ax.transAxes)
        return
    
    # Transformation biexponentielle
    x_trans = biex_transform(x_plot)
    y_trans = biex_transform(y_plot)
    
    # Plot pseudocolor (density)
    try:
        h = ax.hist2d(x_trans, y_trans, bins=100, cmap=cmap, norm=LogNorm(), 
                     cmin=1, rasterized=True)
    except:
        ax.scatter(x_trans, y_trans, s=1, c='blue', alpha=0.3, rasterized=True)
    
    # Gate rectangulaire
    if gate_coords:
        x1, y1, x2, y2 = gate_coords
        x1_t, x2_t = biex_transform([x1, x2])
        y1_t, y2_t = biex_transform([y1, y2])
        rect = patches.Rectangle((x1_t, y1_t), x2_t-x1_t, y2_t-y1_t, 
                                  linewidth=1.5, edgecolor='black', facecolor='none')
        ax.add_patch(rect)
        
        if gate_label and gate_pct is not None:
            ax.text(x2_t, y2_t, f'{gate_label}\n{gate_pct:.1f}%', 
                   fontsize=8, va='bottom', ha='left')
    
    # Quadrants
    if quadrant_coords:
        x_mid, y_mid = quadrant_coords
        x_mid_t = biex_transform([x_mid])[0]
        y_mid_t = biex_transform([y_mid])[0]
        
        xlim = ax.get_xlim() if ax.get_xlim()[1] > ax.get_xlim()[0] else (x_trans.min(), x_trans.max())
        ylim = ax.get_ylim() if ax.get_ylim()[1] > ax.get_ylim()[0] else (y_trans.min(), y_trans.max())
        
        ax.axhline(y=y_mid_t, color='black', linewidth=1, linestyle='-')
        ax.axvline(x=x_mid_t, color='black', linewidth=1, linestyle='-')
        
        if quadrant_labels and quadrant_pcts:
            # Q1: haut-droite, Q2: haut-gauche, Q3: bas-gauche, Q4: bas-droite
            positions = [
                (0.95, 0.95, 'right', 'top'),      # Q1: haut-droite
                (0.05, 0.95, 'left', 'top'),       # Q2: haut-gauche  
                (0.05, 0.05, 'left', 'bottom'),    # Q3: bas-gauche
                (0.95, 0.05, 'right', 'bottom'),   # Q4: bas-droite
            ]
            for i, (label, pct) in enumerate(zip(quadrant_labels, quadrant_pcts)):
                if i < len(positions):
                    px, py, ha, va = positions[i]
                    ax.text(px, py, f'{label}\n{pct:.1f}%', transform=ax.transAxes,
                           fontsize=7, ha=ha, va=va, 
                           bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
    
    ax.set_xlabel(x_label, fontsize=9)
    ax.set_ylabel(y_label, fontsize=9)
    if title:
        ax.set_title(title, fontsize=10, fontweight='bold')


def create_histogram(ax, data, channel, marker, gate_value=None, gate_label=None, gate_pct=None):
    """Crée un histogramme style FlowJo"""
    valid = np.isfinite(data) & (data > 0)
    plot_data = biex_transform(data[valid])
    
    ax.hist(plot_data, bins=100, color='lightgray', edgecolor='darkgray', linewidth=0.5)
    ax.fill_between(ax.get_xlim(), 0, ax.get_ylim()[1] if ax.get_ylim()[1] > 0 else 100, alpha=0.3, color='lightgray')
    
    if gate_value is not None:
        gate_t = biex_transform([gate_value])[0]
        ax.axvline(x=gate_t, color='black', linewidth=1.5)
        
        if gate_label and gate_pct is not None:
            ax.text(0.95, 0.95, f'{gate_label}\n{gate_pct:.1f}%', transform=ax.transAxes,
                   fontsize=8, ha='right', va='top',
                   bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
    
    ax.set_xlabel(f'{marker} ({channel})', fontsize=9)
    ax.set_ylabel('Count', fontsize=9)
    ax.set_title(marker, fontsize=10, fontweight='bold')


def run_flowjo_workflow(data, reader):
    """Exécute le workflow FlowJo complet"""
    
    results = {'gates': {}, 'stats': [], 'populations': {}}
    n_total = len(data)
    
    # ==================== ÉTAPE 1: CELLS ====================
    fsc_a = find_channel(data, ['FSC-A'])
    ssc_a = find_channel(data, ['SSC-A'])
    
    if fsc_a and ssc_a:
        # Gate Cells (exclure débris - FSC et SSC bas)
        fsc_thresh = np.percentile(data[fsc_a], 5)
        ssc_thresh = np.percentile(data[ssc_a], 5)
        fsc_max = np.percentile(data[fsc_a], 99)
        ssc_max = np.percentile(data[ssc_a], 99)
        
        cells_gate = (data[fsc_a] > fsc_thresh) & (data[ssc_a] > ssc_thresh) & \
                     (data[fsc_a] < fsc_max) & (data[ssc_a] < ssc_max)
        
        results['gates']['Cells'] = cells_gate
        results['populations']['Cells'] = data[cells_gate]
        n_cells = cells_gate.sum()
        results['stats'].append({
            'Population': 'Cells', 'Parent': 'Ungated', 
            'Count': n_cells, '% Parent': round(n_cells/n_total*100, 1),
            'Coords': (fsc_thresh, ssc_thresh, fsc_max, ssc_max)
        })
    else:
        cells_gate = pd.Series(True, index=data.index)
        results['gates']['Cells'] = cells_gate
        n_cells = len(data)
    
    # ==================== ÉTAPE 2: SINGLE CELLS ====================
    fsc_h = find_channel(data, ['FSC-H'])
    
    if fsc_a and fsc_h:
        cells_data = data[cells_gate]
        
        # Gate singlets (ratio FSC-A/FSC-H)
        ratio = cells_data[fsc_h] / (cells_data[fsc_a] + 1)
        median_ratio = ratio.median()
        mad = (ratio - median_ratio).abs().median()
        
        singlets_mask = (ratio > median_ratio - 2*mad) & (ratio < median_ratio + 2*mad)
        
        singlets_gate = pd.Series(False, index=data.index)
        singlets_gate.loc[cells_gate] = singlets_mask.values
        
        results['gates']['Single Cells'] = singlets_gate
        results['populations']['Single Cells'] = data[singlets_gate]
        n_singlets = singlets_gate.sum()
        results['stats'].append({
            'Population': 'Single Cells', 'Parent': 'Cells',
            'Count': n_singlets, '% Parent': round(n_singlets/n_cells*100, 1) if n_cells > 0 else 0
        })
    else:
        singlets_gate = cells_gate.copy()
        results['gates']['Single Cells'] = singlets_gate
        n_singlets = singlets_gate.sum()
    
    # ==================== ÉTAPE 3: LIVE ====================
    livedead = find_channel(data, ['LiveDead', 'Viab', 'Live'])
    
    if livedead:
        singlets_data = data[singlets_gate]
        
        # Live = LiveDead négatif (bas)
        ld_thresh = np.percentile(singlets_data[livedead], 85)  # 85% sont vivantes typiquement
        
        live_mask = singlets_data[livedead] < ld_thresh
        
        live_gate = pd.Series(False, index=data.index)
        live_gate.loc[singlets_gate] = live_mask.values
        
        results['gates']['Live'] = live_gate
        results['populations']['Live'] = data[live_gate]
        n_live = live_gate.sum()
        results['stats'].append({
            'Population': 'Live', 'Parent': 'Single Cells',
            'Count': n_live, '% Parent': round(n_live/n_singlets*100, 1) if n_singlets > 0 else 0,
            'Threshold': ld_thresh
        })
    else:
        live_gate = singlets_gate.copy()
        results['gates']['Live'] = live_gate
        n_live = live_gate.sum()
    
    # ==================== ÉTAPE 4: hCD45+ (Leucocytes humains) ====================
    hcd45 = find_channel(data, ['PerCP-A', 'hCD45'])
    mcd45 = find_channel(data, ['APC-Fire750', 'mCD45'])
    
    if hcd45:
        live_data = data[live_gate]
        
        # hCD45 positif
        hcd45_thresh = np.percentile(live_data[hcd45], 15)
        
        if mcd45:
            # hCD45+ et mCD45-
            mcd45_thresh = np.percentile(live_data[mcd45], 90)
            leuco_mask = (live_data[hcd45] > hcd45_thresh) & (live_data[mcd45] < mcd45_thresh)
        else:
            leuco_mask = live_data[hcd45] > hcd45_thresh
        
        leuco_gate = pd.Series(False, index=data.index)
        leuco_gate.loc[live_gate] = leuco_mask.values
        
        results['gates']['hCD45+'] = leuco_gate
        results['populations']['hCD45+'] = data[leuco_gate]
        n_leuco = leuco_gate.sum()
        results['stats'].append({
            'Population': 'hCD45+ (Leucocytes)', 'Parent': 'Live',
            'Count': n_leuco, '% Parent': round(n_leuco/n_live*100, 1) if n_live > 0 else 0,
            'Threshold': hcd45_thresh
        })
    else:
        leuco_gate = live_gate.copy()
        results['gates']['hCD45+'] = leuco_gate
        n_leuco = leuco_gate.sum()
    
    # ==================== ÉTAPE 5: SOUS-POPULATIONS ====================
    leuco_data = data[leuco_gate]
    
    # 5a. NK cells (CD56 vs CD16)
    cd56 = find_channel(data, ['PE-Cy7', 'CD56'])
    cd16 = find_channel(data, ['NovaFluor', 'CD16'])
    
    if cd56 and cd16 and len(leuco_data) > 0:
        cd56_thresh = np.percentile(leuco_data[cd56], 60)
        cd16_thresh = np.percentile(leuco_data[cd16], 60)
        
        nk_mask = (leuco_data[cd56] > cd56_thresh) | (leuco_data[cd16] > cd16_thresh)
        nk_gate = pd.Series(False, index=data.index)
        nk_gate.loc[leuco_gate] = nk_mask.values
        
        results['gates']['NK cells'] = nk_gate
        n_nk = nk_gate.sum()
        results['stats'].append({
            'Population': 'NK cells', 'Parent': 'hCD45+',
            'Count': n_nk, '% Parent': round(n_nk/n_leuco*100, 1) if n_leuco > 0 else 0
        })
        
        # CD56+CD16+ quadrant
        results['quadrants_nk'] = {
            'thresholds': (cd56_thresh, cd16_thresh),
            'labels': ['CD56+CD16+', 'CD56+CD16-', 'CD56-CD16-', 'CD56-CD16+']
        }
    
    # 5b. B cells vs T cells (CD19 vs CD3)
    cd19 = find_channel(data, ['PE-Fire700', 'CD19'])
    cd3 = find_channel(data, ['AF488', 'CD3'])
    
    if cd19 and cd3 and len(leuco_data) > 0:
        cd19_thresh = np.percentile(leuco_data[cd19], 70)
        cd3_thresh = np.percentile(leuco_data[cd3], 40)
        
        # B cells: CD19+ CD3-
        b_mask = (leuco_data[cd19] > cd19_thresh) & (leuco_data[cd3] < cd3_thresh)
        b_gate = pd.Series(False, index=data.index)
        b_gate.loc[leuco_gate] = b_mask.values
        results['gates']['B cells'] = b_gate
        n_b = b_gate.sum()
        
        # T cells: CD19- CD3+
        t_mask = (leuco_data[cd19] < cd19_thresh) & (leuco_data[cd3] > cd3_thresh)
        t_gate = pd.Series(False, index=data.index)
        t_gate.loc[leuco_gate] = t_mask.values
        results['gates']['T cells'] = t_gate
        results['populations']['T cells'] = data[t_gate]
        n_t = t_gate.sum()
        
        results['stats'].append({
            'Population': 'B cells', 'Parent': 'hCD45+',
            'Count': n_b, '% Parent': round(n_b/n_leuco*100, 1) if n_leuco > 0 else 0
        })
        results['stats'].append({
            'Population': 'T cells', 'Parent': 'hCD45+',
            'Count': n_t, '% Parent': round(n_t/n_leuco*100, 1) if n_leuco > 0 else 0
        })
        
        results['quadrants_bt'] = {
            'thresholds': (cd3_thresh, cd19_thresh),
            'labels': ['B cells', 'DP', 'T cells', 'CD19-CD3-']
        }
    
    # 5c. CD4 vs CD8 (sur T cells)
    cd4 = find_channel(data, ['BV650', 'CD4'])
    cd8 = find_channel(data, ['BUV805', 'CD8'])
    
    if cd4 and cd8 and 'T cells' in results['gates']:
        t_data = data[results['gates']['T cells']]
        n_t = len(t_data)
        
        if n_t > 0:
            cd4_thresh = np.percentile(t_data[cd4], 30)
            cd8_thresh = np.percentile(t_data[cd8], 70)
            
            # CD4+ T cells
            cd4_mask = (t_data[cd4] > cd4_thresh) & (t_data[cd8] < cd8_thresh)
            cd4_gate = pd.Series(False, index=data.index)
            cd4_gate.loc[results['gates']['T cells']] = cd4_mask.values
            results['gates']['CD4+ T cells'] = cd4_gate
            results['populations']['CD4+ T cells'] = data[cd4_gate]
            n_cd4 = cd4_gate.sum()
            
            # CD8+ T cells
            cd8_mask = (t_data[cd4] < cd4_thresh) & (t_data[cd8] > cd8_thresh)
            cd8_gate = pd.Series(False, index=data.index)
            cd8_gate.loc[results['gates']['T cells']] = cd8_mask.values
            results['gates']['CD8+ T cells'] = cd8_gate
            n_cd8 = cd8_gate.sum()
            
            results['stats'].append({
                'Population': 'CD4+ T cells', 'Parent': 'T cells',
                'Count': n_cd4, '% Parent': round(n_cd4/n_t*100, 1) if n_t > 0 else 0
            })
            results['stats'].append({
                'Population': 'CD8+ T cells', 'Parent': 'T cells',
                'Count': n_cd8, '% Parent': round(n_cd8/n_t*100, 1) if n_t > 0 else 0
            })
            
            results['quadrants_cd4cd8'] = {
                'thresholds': (cd4_thresh, cd8_thresh),
                'labels': ['DP', 'CD8+', 'DN', 'CD4+']
            }
    
    # 5d. Treg (FoxP3 vs CD25) sur CD4+ T cells
    foxp3 = find_channel(data, ['eFluor450', 'FoxP3'])
    cd25 = find_channel(data, ['BV785', 'CD25'])
    
    if foxp3 and cd25 and 'CD4+ T cells' in results['gates']:
        cd4_data = data[results['gates']['CD4+ T cells']]
        n_cd4 = len(cd4_data)
        
        if n_cd4 > 0:
            foxp3_thresh = np.percentile(cd4_data[foxp3], 90)
            cd25_thresh = np.percentile(cd4_data[cd25], 85)
            
            treg_mask = (cd4_data[foxp3] > foxp3_thresh) & (cd4_data[cd25] > cd25_thresh)
            treg_gate = pd.Series(False, index=data.index)
            treg_gate.loc[results['gates']['CD4+ T cells']] = treg_mask.values
            results['gates']['Treg'] = treg_gate
            n_treg = treg_gate.sum()
            
            results['stats'].append({
                'Population': 'Treg (FoxP3+CD25+)', 'Parent': 'CD4+ T cells',
                'Count': n_treg, '% Parent': round(n_treg/n_cd4*100, 1) if n_cd4 > 0 else 0
            })
            
            results['quadrants_treg'] = {
                'thresholds': (foxp3_thresh, cd25_thresh),
                'labels': ['Treg', 'CD25+', 'FoxP3-CD25-', 'FoxP3+']
            }
    
    # Marqueurs fonctionnels pour histogrammes
    results['functional_markers'] = {
        'PDL1': find_channel(data, ['BUV395', 'PDL1']),
        'CD161': find_channel(data, ['BV711', 'CD161']),
        'PD1': find_channel(data, ['RB780', 'PD1']),
        'CD107a': find_channel(data, ['PerCP-Vio700', 'CD107a']),
        'Granzyme B': find_channel(data, ['PE-Dazzle', 'Granzyme']),
    }
    
    return results


def create_flowjo_figure(data, reader, results, xlims=None, ylims=None):
    """Crée la figure complète style FlowJo"""
    
    fig = plt.figure(figsize=(20, 16), dpi=100)
    
    # Définir la grille (similaire à l'image FlowJo)
    # Ligne 1: 6 plots (Cells, Single Cells, Live, hCD45, etc.)
    # Ligne 2: 6 plots (sous-populations avec quadrants)
    # Ligne 3: 6 histogrammes
    
    n_total = len(data)
    
    # ===== LIGNE 1: GATING PRINCIPAL =====
    
    # Plot 1: FSC-A vs SSC-A (Cells)
    ax1 = fig.add_subplot(3, 6, 1)
    fsc_a = find_channel(data, ['FSC-A'])
    ssc_a = find_channel(data, ['SSC-A'])
    if fsc_a and ssc_a:
        gate_info = next((s for s in results['stats'] if s['Population'] == 'Cells'), None)
        coords = gate_info.get('Coords') if gate_info else None
        pct = gate_info['% Parent'] if gate_info else 0
        create_flowjo_plot(ax1, data[fsc_a].values, data[ssc_a].values,
                          'FSC-A', 'SSC-A', f'{reader.filename}\nUngated\n{n_total}',
                          gate_coords=coords, gate_label='Cells', gate_pct=pct)
    
    # Plot 2: FSC-A vs FSC-H (Single Cells)
    ax2 = fig.add_subplot(3, 6, 2)
    fsc_h = find_channel(data, ['FSC-H'])
    if fsc_a and fsc_h and 'Cells' in results['gates']:
        cells_data = data[results['gates']['Cells']]
        gate_info = next((s for s in results['stats'] if s['Population'] == 'Single Cells'), None)
        pct = gate_info['% Parent'] if gate_info else 0
        n_cells = results['gates']['Cells'].sum()
        create_flowjo_plot(ax2, cells_data[fsc_a].values, cells_data[fsc_h].values,
                          'FSC-A', 'FSC-H', f'Cells\n{n_cells}',
                          gate_label='Single Cells', gate_pct=pct)
    
    # Plot 3: Live/Dead vs SSC-A (Live)
    ax3 = fig.add_subplot(3, 6, 3)
    livedead = find_channel(data, ['LiveDead', 'Viab'])
    if livedead and ssc_a and 'Single Cells' in results['gates']:
        singlets_data = data[results['gates']['Single Cells']]
        gate_info = next((s for s in results['stats'] if s['Population'] == 'Live'), None)
        pct = gate_info['% Parent'] if gate_info else 0
        n_sing = results['gates']['Single Cells'].sum()
        create_flowjo_plot(ax3, singlets_data[livedead].values, singlets_data[ssc_a].values,
                          'Live/Dead', 'SSC-A', f'Single Cells\n{n_sing}',
                          gate_label='Live', gate_pct=pct)
    
    # Plot 4: hCD45 vs mCD45
    ax4 = fig.add_subplot(3, 6, 4)
    hcd45 = find_channel(data, ['PerCP-A'])
    mcd45 = find_channel(data, ['APC-Fire750'])
    if hcd45 and 'Live' in results['gates']:
        live_data = data[results['gates']['Live']]
        gate_info = next((s for s in results['stats'] if 'hCD45' in s['Population']), None)
        pct = gate_info['% Parent'] if gate_info else 0
        n_live = results['gates']['Live'].sum()
        y_ch = mcd45 if mcd45 else ssc_a
        create_flowjo_plot(ax4, live_data[hcd45].values, live_data[y_ch].values,
                          'hCD45 (PerCP)', 'mCD45' if mcd45 else 'SSC-A', f'Live\n{n_live}',
                          gate_label='hCD45+', gate_pct=pct)
    
    # Plot 5: HLA-ABC
    ax5 = fig.add_subplot(3, 6, 5)
    hla = find_channel(data, ['APC-A', 'HLA'])
    if hla and hcd45 and 'hCD45+' in results['gates']:
        leuco_data = data[results['gates']['hCD45+']]
        n_leuco = len(leuco_data)
        create_flowjo_plot(ax5, leuco_data[hcd45].values, leuco_data[hla].values,
                          'hCD45', 'HLA-ABC', f'hCD45+\n{n_leuco}')
    
    # Plot 6: Leucocytes overview
    ax6 = fig.add_subplot(3, 6, 6)
    if hcd45 and 'hCD45+' in results['gates']:
        leuco_data = data[results['gates']['hCD45+']]
        n_leuco = len(leuco_data)
        create_flowjo_plot(ax6, leuco_data[hcd45].values, leuco_data[ssc_a].values if ssc_a else leuco_data[hcd45].values,
                          'hCD45', 'SSC-A', f'Leucocytes\n{n_leuco}')
    
    # ===== LIGNE 2: SOUS-POPULATIONS AVEC QUADRANTS =====
    
    # Plot 7: NK cells (CD56 vs CD16)
    ax7 = fig.add_subplot(3, 6, 7)
    cd56 = find_channel(data, ['PE-Cy7', 'CD56'])
    cd16 = find_channel(data, ['NovaFluor', 'CD16'])
    if cd56 and cd16 and 'hCD45+' in results['gates']:
        leuco_data = data[results['gates']['hCD45+']]
        n_leuco = len(leuco_data)
        quad = results.get('quadrants_nk', {})
        thresholds = quad.get('thresholds', (0, 0))
        
        # Calculer pourcentages des quadrants
        cd56_t, cd16_t = thresholds
        q_pcts = []
        for cond in [(True, True), (True, False), (False, False), (False, True)]:
            cd56_cond = leuco_data[cd56] > cd56_t if cond[0] else leuco_data[cd56] <= cd56_t
            cd16_cond = leuco_data[cd16] > cd16_t if cond[1] else leuco_data[cd16] <= cd16_t
            q_pcts.append((cd56_cond & cd16_cond).sum() / n_leuco * 100 if n_leuco > 0 else 0)
        
        create_flowjo_plot(ax7, leuco_data[cd16].values, leuco_data[cd56].values,
                          'CD16', 'CD56', f'hCD45+\n{n_leuco}',
                          quadrant_coords=thresholds[::-1],
                          quadrant_labels=['NK cells', 'CD56+', 'DN', 'CD16+'],
                          quadrant_pcts=q_pcts)
    
    # Plot 8: B cells vs T cells (CD19 vs CD3)
    ax8 = fig.add_subplot(3, 6, 8)
    cd19 = find_channel(data, ['PE-Fire700', 'CD19'])
    cd3 = find_channel(data, ['AF488', 'CD3'])
    if cd19 and cd3 and 'hCD45+' in results['gates']:
        leuco_data = data[results['gates']['hCD45+']]
        n_leuco = len(leuco_data)
        quad = results.get('quadrants_bt', {})
        thresholds = quad.get('thresholds', (0, 0))
        
        cd3_t, cd19_t = thresholds
        q_pcts = []
        for cond in [(True, True), (False, True), (False, False), (True, False)]:
            cd3_cond = leuco_data[cd3] > cd3_t if cond[0] else leuco_data[cd3] <= cd3_t
            cd19_cond = leuco_data[cd19] > cd19_t if cond[1] else leuco_data[cd19] <= cd19_t
            q_pcts.append((cd3_cond & cd19_cond).sum() / n_leuco * 100 if n_leuco > 0 else 0)
        
        create_flowjo_plot(ax8, leuco_data[cd3].values, leuco_data[cd19].values,
                          'CD3', 'CD19', f'hCD45+\n{n_leuco}',
                          quadrant_coords=thresholds,
                          quadrant_labels=['B cells', 'DP', 'CD19-CD3-', 'T cells'],
                          quadrant_pcts=q_pcts)
    
    # Plot 9: CD4 vs CD8
    ax9 = fig.add_subplot(3, 6, 9)
    cd4 = find_channel(data, ['BV650', 'CD4'])
    cd8 = find_channel(data, ['BUV805', 'CD8'])
    if cd4 and cd8 and 'T cells' in results['gates']:
        t_data = data[results['gates']['T cells']]
        n_t = len(t_data)
        quad = results.get('quadrants_cd4cd8', {})
        thresholds = quad.get('thresholds', (0, 0))
        
        cd4_t, cd8_t = thresholds
        q_pcts = []
        for cond in [(True, True), (False, True), (False, False), (True, False)]:
            cd4_cond = t_data[cd4] > cd4_t if cond[0] else t_data[cd4] <= cd4_t
            cd8_cond = t_data[cd8] > cd8_t if cond[1] else t_data[cd8] <= cd8_t
            q_pcts.append((cd4_cond & cd8_cond).sum() / n_t * 100 if n_t > 0 else 0)
        
        create_flowjo_plot(ax9, t_data[cd4].values, t_data[cd8].values,
                          'CD4', 'CD8', f'T cells\n{n_t}',
                          quadrant_coords=thresholds,
                          quadrant_labels=['DP', 'CD8+', 'DN', 'CD4+'],
                          quadrant_pcts=q_pcts)
    
    # Plot 10: Treg (FoxP3 vs CD25)
    ax10 = fig.add_subplot(3, 6, 10)
    foxp3 = find_channel(data, ['eFluor450', 'FoxP3'])
    cd25 = find_channel(data, ['BV785', 'CD25'])
    if foxp3 and cd25 and 'CD4+ T cells' in results['gates']:
        cd4_data = data[results['gates']['CD4+ T cells']]
        n_cd4 = len(cd4_data)
        quad = results.get('quadrants_treg', {})
        thresholds = quad.get('thresholds', (0, 0))
        
        foxp3_t, cd25_t = thresholds
        q_pcts = []
        for cond in [(True, True), (False, True), (False, False), (True, False)]:
            f_cond = cd4_data[foxp3] > foxp3_t if cond[0] else cd4_data[foxp3] <= foxp3_t
            c_cond = cd4_data[cd25] > cd25_t if cond[1] else cd4_data[cd25] <= cd25_t
            q_pcts.append((f_cond & c_cond).sum() / n_cd4 * 100 if n_cd4 > 0 else 0)
        
        create_flowjo_plot(ax10, cd4_data[foxp3].values, cd4_data[cd25].values,
                          'FoxP3', 'CD25', f'CD4+ T cells\n{n_cd4}',
                          quadrant_coords=thresholds,
                          quadrant_labels=['Treg', 'CD25+', 'DN', 'FoxP3+'],
                          quadrant_pcts=q_pcts)
    
    # Plots 11-12: Espaces pour autres populations
    ax11 = fig.add_subplot(3, 6, 11)
    ax11.axis('off')
    ax12 = fig.add_subplot(3, 6, 12)
    ax12.axis('off')
    
    # ===== LIGNE 3: HISTOGRAMMES =====
    functional = results.get('functional_markers', {})
    hist_population = 'B cells' if 'B cells' in results['gates'] else 'hCD45+'
    hist_data = data[results['gates'].get(hist_population, results['gates'].get('hCD45+', pd.Series(True, index=data.index)))]
    n_hist = len(hist_data)
    
    hist_idx = 13
    for marker_name, channel in functional.items():
        if channel and hist_idx <= 18:
            ax = fig.add_subplot(3, 6, hist_idx)
            
            if channel in hist_data.columns:
                marker_data = hist_data[channel].values
                valid = np.isfinite(marker_data) & (marker_data > 0)
                
                if valid.sum() > 0:
                    # Calculer % positif
                    thresh = np.percentile(marker_data[valid], 85)
                    pct_pos = (marker_data > thresh).sum() / len(marker_data) * 100
                    
                    create_histogram(ax, marker_data, channel, marker_name,
                                   gate_value=thresh, gate_label=f'{marker_name}+', gate_pct=pct_pos)
                    ax.set_title(f'{hist_population}\n{n_hist}', fontsize=9)
            
            hist_idx += 1
    
    # Cacher les axes non utilisés
    for i in range(hist_idx, 19):
        ax = fig.add_subplot(3, 6, i)
        ax.axis('off')
    
    plt.tight_layout()
    return fig


def export_flowjo_excel(reader, results):
    """Export Excel style FlowJo"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except:
        return None
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", fill_type="solid")
    
    # Feuille Statistiques
    ws = wb.create_sheet("Population Statistics")
    headers = ['Population', 'Parent', 'Count', '% of Parent', '% of Total']
    ws.append(headers)
    
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
    
    n_total = len(reader.data)
    for stat in results['stats']:
        ws.append([
            stat['Population'],
            stat['Parent'],
            stat['Count'],
            f"{stat['% Parent']}%",
            f"{round(stat['Count']/n_total*100, 2)}%"
        ])
    
    # Feuille Marqueurs
    ws2 = wb.create_sheet("Marker Summary")
    ws2.append(['Marker', 'Channel', 'Population', '% Positive'])
    
    for c_idx in range(1, 5):
        cell = ws2.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
    
    # Ajuster largeurs
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==================== INTERFACE STREAMLIT ====================

st.markdown('<h1 class="main-header">🔬 FACS Analysis - FlowJo Style</h1>', unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.markdown("### ⚙️ Options")
    auto_gate = st.checkbox("Gating automatique", value=True)
    show_stats = st.checkbox("Afficher statistiques", value=True)

# Upload
uploaded_file = st.file_uploader("📁 Fichier FCS", type=['fcs'])

if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix='.fcs') as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name
    
    try:
        with st.spinner("Chargement..."):
            reader = FCSReader(tmp_path)
        
        col1, col2, col3 = st.columns(3)
        col1.metric("Événements", f"{len(reader.data):,}")
        col2.metric("Canaux", len(reader.channels))
        col3.metric("Fichier", reader.filename[:20])
        
        if st.button("🚀 Analyser (Workflow FlowJo)", type="primary", use_container_width=True):
            with st.spinner("Analyse du workflow..."):
                results = run_flowjo_workflow(reader.data, reader)
                st.session_state.results = results
                st.session_state.reader = reader
        
        if 'results' in st.session_state:
            results = st.session_state.results
            reader = st.session_state.reader
            
            # Statistiques
            if show_stats:
                st.markdown("### 📊 Populations")
                stats_df = pd.DataFrame(results['stats'])
                st.dataframe(stats_df, use_container_width=True)
            
            # Figure FlowJo
            st.markdown("### 🎨 Visualisation FlowJo")
            
            with st.spinner("Génération des graphiques..."):
                fig = create_flowjo_figure(reader.data, reader, results)
                st.pyplot(fig)
                st.session_state.fig = fig
            
            # Export
            col1, col2 = st.columns(2)
            with col1:
                if st.button("📥 Export PNG (300 DPI)"):
                    buf = io.BytesIO()
                    st.session_state.fig.savefig(buf, format='png', dpi=300, bbox_inches='tight')
                    buf.seek(0)
                    st.download_button("Télécharger", buf, f"{reader.filename}_flowjo.png", "image/png")
            
            with col2:
                if st.button("📥 Export Excel"):
                    excel = export_flowjo_excel(reader, results)
                    if excel:
                        st.download_button("Télécharger", excel, f"{reader.filename}_stats.xlsx")
    
    except Exception as e:
        st.error(f"Erreur: {e}")
        st.exception(e)

st.markdown("---")
st.markdown("<p style='text-align:center;color:gray;'>🔬 FACS FlowJo Style | Workflow Immunophénotypage</p>", unsafe_allow_html=True)
